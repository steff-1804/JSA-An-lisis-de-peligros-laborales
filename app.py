from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from PIL import Image
import json
import base64
import os

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/exportar_excel", methods=["POST"])
def exportar_excel():
    data = json.loads(request.form.get("data", "[]"))
    info = json.loads(request.form.get("info_general", "{}"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz JSA"

    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    hazard_fill = PatternFill("solid", fgColor="FCE4D6")
    existing_fill = PatternFill("solid", fgColor="DAEEF3")
    recommended_fill = PatternFill("solid", fgColor="CCFF66")
    post_fill = PatternFill("solid", fgColor="D9EAD3")

    green_fill = PatternFill("solid", fgColor="86EFAC")
    yellow_fill = PatternFill("solid", fgColor="FDE047")
    red_fill = PatternFill("solid", fgColor="F87171")

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    ws["A1"] = "Risk Assessment / Job Hazard Analysis Evaluación de Riesgos / Análisis de Peligros Laborales"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A2:B5")
    ws["A2"] = f"TASK DESCRIPTION: DESCRIPCIÓN DEL TRABAJO:\n{info.get('descripcion_trabajo', '')}"
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(vertical="top", wrap_text=True)

    ws.merge_cells("C2:K2")
    ws["C2"] = f"TITLE: TÍTULO DEL JSA: {info.get('titulo_jsa', '')}"
    ws["C2"].font = Font(bold=True)

    ws.merge_cells("C3:K3")
    ws["C3"] = f"DOCUMENT REFERENCE NUMBER: {info.get('documento_referencia', '')}"
    ws["C3"].font = Font(bold=True)

    ws.merge_cells("C4:K4")
    ws["C4"] = f"TEAM MEMBERS: MIEMBROS DEL EQUIPO JSA: {info.get('miembros_equipo', '')}"
    ws["C4"].font = Font(bold=True)

    ws.merge_cells("C5:K5")
    ws["C5"] = f"DATE: {info.get('fecha_jsa', '')}"
    ws["C5"].font = Font(bold=True)

    ws.merge_cells("L1:M1")
    ws["L1"] = "JSA"
    ws["L1"].font = Font(bold=True, size=13)
    ws["L1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("L2:M5")

    logo_path = os.path.join(app.root_path, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        try:
            logo = ExcelImage(logo_path)
            logo.width = 120
            logo.height = 90
            ws.add_image(logo, "L2")
        except Exception:
            ws["L2"] = "LOGO"

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=13):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

    ws.merge_cells("A6:B6")
    ws["A6"] = "Job / Activity Step"

    ws["C6"] = "Photos"

    ws.merge_cells("D6:E6")
    ws["D6"] = "Hazards and Potential Consequences"

    ws.merge_cells("F6:J6")
    ws["F6"] = "EXISTING / CURRENT CONDITION RANKING"

    ws.merge_cells("K6:K7")
    ws["K6"] = "Recommended Controls"

    ws.merge_cells("L6:M6")
    ws["L6"] = "POST RANKINGS"

    section_cells = ["A6", "C6", "D6", "F6", "K6", "L6"]
    for cell in section_cells:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[cell].border = border

    ws["A6"].fill = subheader_fill
    ws["C6"].fill = subheader_fill
    ws["D6"].fill = hazard_fill
    ws["F6"].fill = existing_fill
    ws["K6"].fill = recommended_fill
    ws["L6"].fill = post_fill

    headers = [
        "Área",
        "Job / Activity Step",
        "Photos",
        "Type",
        "Hazards and Potential Consequences",
        "SEV 1-5",
        "LIKL 1-5",
        "Type",
        "Existing Controls",
        "CONT 1-5",
        "RPN",
        "Type",
        "Recommended Controls",
        "SEV 1-5",
        "LIKL 1-5",
        "CONT 1-5",
        "RPN"
    ]

    start_row = 7
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = subheader_fill
        cell.border = border

    data_start = 8

    for row_index, item in enumerate(data, start=data_start):
        sev = int(item.get("sev", 1))
        likl = int(item.get("likl", 1))
        cont = int(item.get("cont", 1))

        sev_post = int(item.get("sev_post", 1))
        likl_post = int(item.get("likl_post", 1))
        cont_post = int(item.get("cont_post", 1))

        rpn = sev * likl * cont
        rpn_post = sev_post * likl_post * cont_post

        values = [
            item.get("area", ""),
            item.get("actividad", ""),
            "",
            item.get("tipo_riesgo", ""),
            item.get("peligro", ""),
            sev,
            likl,
            item.get("tipo_control_existente", ""),
            item.get("existing_controls", ""),
            cont,
            rpn,
            item.get("tipo_control_recomendado", ""),
            item.get("recommended_controls", ""),
            sev_post,
            likl_post,
            cont_post,
            rpn_post
        ]

        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col in [11, 17]:
            cell = ws.cell(row=row_index, column=col)

            if cell.value <= 10:
                cell.fill = green_fill
            elif cell.value <= 30:
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        foto_base64 = item.get("foto", "")
        if foto_base64:
            try:
                _, encoded = foto_base64.split(",", 1)
                image_data = base64.b64decode(encoded)

                image_stream = BytesIO(image_data)
                pil_image = Image.open(image_stream)
                pil_image.thumbnail((150, 100))

                final_stream = BytesIO()
                pil_image.save(final_stream, format="PNG")
                final_stream.seek(0)

                excel_img = ExcelImage(final_stream)
                excel_img.width = 140
                excel_img.height = 95

                ws.add_image(excel_img, f"C{row_index}")
                ws.row_dimensions[row_index].height = 80

            except Exception:
                ws.cell(row=row_index, column=3).value = "Imagen no válida"

    column_widths = {
        "A": 18,
        "B": 30,
        "C": 22,
        "D": 10,
        "E": 45,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 45,
        "J": 10,
        "K": 10,
        "L": 10,
        "M": 45,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A8"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Matriz_JSA.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=False)
